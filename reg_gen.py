from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader, select_autoescape
import datetime
import copy
import os
import sys
import fnmatch
import argparse

class Field:
  name = ""
  description = ""
  bit_offset = 0
  reset_val = 0
  bit_width = 1
  access = "rw_hr"

  def __init__(self, name, description, bit_offset, reset_val, bit_width, access):
    self.name = name
    self.description = description
    self.bit_offset = bit_offset
    self.reset_val = reset_val
    self.bit_width = bit_width
    self.access = access

class Register:
  name = ""
  description = ""
  addr_offset = 0
  size = 32  # check width TODO
  fields = list()

  def __init__(self, name, description, addr_offset, size, fields):
    self.name = name
    self.description = description
    self.addr_offset = addr_offset
    self.size = size
    self.fields = fields


#################################################################
# Description    : Environment                                  #
#################################################################
env = Environment(
    loader=FileSystemLoader("templates"),
    autoescape=select_autoescape(
        disabled_extensions=(".txt", ".sv", ".v"),
        default=False
    )
)


def reg_inputs(reg):
  cur_index = 0
  reserved_count = 0
  temp_output = "}"
  # single full size field :
  if reg.fields[0].bit_width == 32:
    return str(reg.name.upper() + "_" + reg.fields[0].name + "_data_in")
  else:
    for field in reg.fields:
      if field.bit_offset > cur_index:
        if temp_output != "}":
          temp_output = ", " + temp_output
        temp_output = reg.name.upper() + "_reserved_" + str(reserved_count) + temp_output
        reserved_count = reserved_count + 1
      cur_index = field.bit_width + field.bit_offset
      if temp_output != "}":
        temp_output = ", " + temp_output
      temp_output = reg.name.upper() + "_" + field.name + "_data_in" + temp_output
    if cur_index < 32:
      temp_output = "{" + reg.name.upper() + "_reserved_" + str(reserved_count) + ", " + temp_output
    else:
      temp_output = "{" + temp_output
    return temp_output

def reg_outputs(reg):
  cur_index = 0
  temp_output = "}"
  # single full size field :
  if reg.fields[0].bit_width == 32:
    return str(reg.name.upper() + "_" + reg.fields[0].name + "_data_out")
  else:
    for field in reg.fields:
      if field.bit_offset > cur_index:
        if temp_output != "}":
          temp_output = ", " + temp_output
        temp_output = str(field.bit_offset - cur_index) + "'b0" + temp_output
      cur_index = field.bit_width + field.bit_offset
      if temp_output != "}":
        temp_output = ", " + temp_output
      if field.access != "wo" and field.access != "w_hrw":
        temp_output = reg.name.upper() + "_" + field.name + "_data_out" + temp_output
      else:
        temp_output = str(cur_index - field.bit_offset) + "'b0" + temp_output
    if cur_index < 32:
      temp_output = "{" + str(32 - cur_index) + "'b0" + ", " + temp_output
    else:
      temp_output = "{" + temp_output
    return temp_output


#################################################################
# Description    :                                              #
#################################################################
if __name__ == '__main__':

    parser = argparse.ArgumentParser(description="Regs gen script",
                                    formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("-x", "--xlsx", help="Path to the *.xlsx file")
    parser.add_argument("-n", "--name", help="Module name")
    args = parser.parse_args()

    if args.xlsx == None:
        print("Add the path to the excel file -x/--xlsx.")
        exit()
    elif not ".xlsx" in args.xlsx:
        print("The argument must have a *.xlsx extension.")
        exit()

    if args.name == None:
        print("Specify the module file name -n/--name (no extension).")
        exit()
    elif "." in args.name:
        print("'.' is not allowed in the name.")
        exit()
    else:
        module_name = args.name

    module_file = ("./"+module_name+".sv")

    NAME                = 1
    DESCRIPTION         = 2
    ADDR_OFFSET         = 3
    SIZE                = 4
    FIELD_NAME          = 5
    FIELD_DES           = 6
    FIELD_BIT_OFFSET    = 7
    FIELD_RESET_VAL     = 8
    FIELD_BIT_WIDTH     = 9
    FIELD_ACCESS        = 10

    wb = load_workbook(args.xlsx)
    sheet = wb.active

    registers   = []
    fields      = []

    cnt = 2

    for i in range (2, sheet.max_row+1):

        temp_field = Field(sheet.cell(i ,FIELD_NAME      ).value,
                           sheet.cell(i ,FIELD_DES       ).value,
                           sheet.cell(i ,FIELD_BIT_OFFSET).value,
                           sheet.cell(i ,FIELD_RESET_VAL ).value,
                           sheet.cell(i ,FIELD_BIT_WIDTH ).value,
                           sheet.cell(i ,FIELD_ACCESS    ).value)
        # TODO check if there is bitfields overlapping
        fields.append(temp_field)

        if i + 1 == sheet.max_row + 1:
            temp_register = Register(sheet.cell(cnt ,NAME).value,
                                     sheet.cell(cnt ,DESCRIPTION).value,
                                     sheet.cell(cnt ,ADDR_OFFSET).value,
                                     sheet.cell(cnt ,SIZE).value,
                                     copy.deepcopy(fields))
        #else:  # TODO
        #    print("U redu ",i," moraju biti popunjene sve vrednosti ili sve vrednosti da budu None")

        # TODO check if there is register address overlapping
            registers.append(copy.deepcopy(temp_register))

        elif sheet.cell(i+1 ,NAME).value != None and sheet.cell(i+1 ,DESCRIPTION).value != None: # TODO
            temp_register = Register(sheet.cell(cnt ,NAME).value,
                                     sheet.cell(cnt ,DESCRIPTION).value,
                                     sheet.cell(cnt ,ADDR_OFFSET).value,
                                     sheet.cell(cnt ,SIZE).value,
                                     copy.deepcopy(fields))
        #else:  # TODO
        #    print("U redu ",i," moraju biti popunjene sve vrednosti ili sve vrednosti da budu None")

        # TODO check if there is register address overlapping
            registers.append(copy.deepcopy(temp_register))
            fields.clear()
            cnt = i + 1

    date = datetime.datetime.now()

    o_width, o_name, o_field_name = [], [], []
    i_width, i_name, i_field_name = [], [], []
    iw0_width, iw0_name, iw0_field_name = [], [], []
    irw_hw_width, irw_hw_name, irw_hw_field_name = [], [], []
    irwhw_width, irwhw_name, irwhw_field_name = [], [], []
    irhr_width, irhr_name, irhr_field_name = [], [], []
    iwhw_width, iwhw_name, iwhw_field_name = [], [], []
    iwo_width, iwo_name, iwo_field_name = [], [], []
    res_name, res_width, res_cnt = [], [], []
    inf_bit_width, inf_fname, inf_rname = [], [], []
    ino_bit_width, ino_fname, ino_rname = [], [], []
    w0_sd_fname, w0_sd_rname, w0_sd_bit_width = [], [], []
    rwhw_sd_fname, rwhw_sd_rname, rwhw_sd_bit_width = [], [], []
    rw_hw_sd_fname, rw_hw_sd_rname, rw_hw_sd_bit_width = [], [], []
    whw_sd_fname, whw_sd_rname, whw_sd_bit_width = [], [], []
    rhr_sd_fname, rhr_sd_rname, rhr_sd_bit_width = [], [], []
    wdec_name       = []
    adec_name       = []
    adec_addr_off   = []
    wen_name        = []
    wdo_name        = []
    a_en_rw_name    = []
    dom_addr_off    = []
    dom_name        = []
    i_rw_name       = []
    i_rw_fbit_w     = []
    i_rw_fname      = []
    i_rw_frstv      = []
    i_rw0_name      = []
    i_rw0_fbit_w    = []
    i_rw0_fname     = []
    i_rw0_frstv     = []
    i_rw_hw_name      = []
    i_rw_hw_fbit_w    = []
    i_rw_hw_fname     = []
    i_rw_hw_frstv     = []
    i_rwhw_name      = []
    i_rwhw_fbit_w    = []
    i_rwhw_fname     = []
    i_rwhw_frstv     = []
    i_ro_name       = []
    i_ro_fname      = []
    i_ro_fbit_w     = []
    i_ro_frstv     = []
    i_rhr_name, i_rhr_fname, i_rhr_fbit_w, i_rhr_frstv = [], [], [], []
    i_whw_name, i_whw_fname, i_whw_fbit_w, i_whw_frstv = [], [], [], []
    i_wo_name, i_wo_fname, i_wo_fbit_w, i_wo_frstv = [], [], [], []
    reg_in          = []
    dos_wr_name     = []
    dos_wr_out      = []
    cout_name       = []
    cout_fname      = []
    rwhwcout_name   = []
    rwhwcout_fname  = []
    rhrcout_name    = []
    rhrcout_fname   = []
    whwcout_name    = []
    whwcout_fname   = []
    wocout_name     = []
    wocout_fname    = []
    ae_name         = []
    ae_nro_name     = []
    o_reg_desc, i_reg_desc, iw0_reg_desc, irwhw_reg_desc         = [], [], [], []
    irhr_reg_desc, iwhw_reg_desc, iwo_reg_desc                   = [], [], []
    o_field_desc, i_field_desc, iw0_field_desc, irwhw_field_desc = [], [], [], []
    irhr_field_desc, iwhw_field_desc, iwo_field_desc             = [], [], []
    irw_hw_reg_desc, irw_hw_field_desc = [], []

    # Add inputs and outputs
    for reg in registers:
        for field in reg.fields:
            if field.access == "rw_hr":
                o_name.append(reg.name.upper())
                o_field_name.append(field.name)
                o_reg_desc.append(reg.description)
                o_field_desc.append(field.description)
                o_width.append(str(field.bit_width))
            elif field.access == "rc_w0":
                iw0_name.append(reg.name.upper())
                iw0_field_name.append(field.name)
                iw0_reg_desc.append(reg.description)
                iw0_field_desc.append(field.description)
                iw0_width.append(str(field.bit_width))
            elif field.access == "rw_hw":
                irw_hw_name.append(reg.name.upper())
                irw_hw_field_name.append(field.name)
                irw_hw_reg_desc.append(reg.description)
                irw_hw_field_desc.append(field.description)
                irw_hw_width.append(str(field.bit_width))
            elif field.access == "rw_hrw":
                irwhw_name.append(reg.name.upper())
                irwhw_field_name.append(field.name)
                irwhw_reg_desc.append(reg.description)
                irwhw_field_desc.append(field.description)
                irwhw_width.append(str(field.bit_width))
            elif field.access == "ro":
                i_name.append(reg.name.upper())
                i_field_name.append(field.name)
                i_reg_desc.append(reg.description)
                i_field_desc.append(field.description)
                i_width.append(str(field.bit_width))
            elif field.access == "r_hrw":
                irhr_name.append(reg.name.upper())
                irhr_field_name.append(field.name)
                irhr_reg_desc.append(reg.description)
                irhr_field_desc.append(field.description)
                irhr_width.append(str(field.bit_width))
            elif field.access == "w_hrw":
                iwhw_name.append(reg.name.upper())
                iwhw_field_name.append(field.name)
                iwhw_reg_desc.append(reg.description)
                iwhw_field_desc.append(field.description)
                iwhw_width.append(str(field.bit_width))
            elif field.access == "wo":
                iwo_name.append(reg.name.upper())
                iwo_field_name.append(field.name)
                iwo_reg_desc.append(reg.description)
                iwo_field_desc.append(field.description)
                iwo_width.append(str(field.bit_width))

    out_t       = list(zip(o_name,      o_width,        o_field_name    , o_reg_desc, o_field_desc))
    in_t        = list(zip(i_name,      i_width,        i_field_name    , i_reg_desc, i_field_desc))
    in_w0_t     = list(zip(iw0_name,    iw0_width,      iw0_field_name  , iw0_reg_desc, iw0_field_desc))
    in_rwhw     = list(zip(irwhw_name,  irwhw_width,    irwhw_field_name, irwhw_reg_desc, irwhw_field_desc))
    in_rhr      = list(zip(irhr_name,   irhr_width,     irhr_field_name , irhr_reg_desc, irhr_field_desc))
    in_hwh      = list(zip(iwhw_name,   iwhw_width,     iwhw_field_name , iwhw_reg_desc, iwhw_field_desc))
    in_wo       = list(zip(iwo_name,    iwo_width,      iwo_field_name  , iwo_reg_desc, iwo_field_desc))
    irw_hw_t    = list(zip(irw_hw_name, irw_hw_width,   irw_hw_field_name,irw_hw_reg_desc, irw_hw_field_desc))

    # Reserved for read and write purposes.
    for reg in registers:
        index   = 0
        reser_cnt = 0
        for field in reg.fields:
            if field.bit_offset > index:
                res_name.append(reg.name.upper())
                res_width.append(str(field.bit_offset - index))
                res_cnt.append(str(reser_cnt))
                reser_cnt = reser_cnt + 1
            index = field.bit_width + field.bit_offset
        if index < 32:
            res_name.append(reg.name.upper())
            res_width.append(str(32 - index))
            res_cnt.append(str(reser_cnt))

    res_t = zip(res_name, res_width, res_cnt)

    # WIRE: Field inputs/outputs
    for reg in registers:
        for field in reg.fields:
            inf_bit_width.append(str(field.bit_width))
            inf_fname.append(field.name)
            inf_rname.append(reg.name.upper())
            ino_bit_width.append(str(field.bit_width))
            ino_fname.append(field.name)
            ino_rname.append(reg.name.upper())

    inf_t = zip(inf_bit_width, inf_fname, inf_rname)
    ino_t = zip(ino_bit_width, ino_fname, ino_rname)
    en_t  = zip(ino_fname, ino_rname)

    # WIRE: sel_data
    for reg in registers:
        for field in reg.fields:
            if field.access == "rc_w0":
                w0_sd_fname.append(field.name)
                w0_sd_rname.append(reg.name.upper())
                w0_sd_bit_width.append(str(field.bit_width))
            if field.access == "rw_hw":
                rw_hw_sd_fname.append(field.name)
                rw_hw_sd_rname.append(reg.name.upper())
                rw_hw_sd_bit_width.append(str(field.bit_width))
            elif field.access == "rw_hrw":
                rwhw_sd_fname.append(field.name)
                rwhw_sd_rname.append(reg.name.upper())
                rwhw_sd_bit_width.append(str(field.bit_width))
            elif field.access == "w_hrw":
                whw_sd_fname.append(field.name)
                whw_sd_rname.append(reg.name.upper())
                whw_sd_bit_width.append(str(field.bit_width))


    w0_sd_t = zip(w0_sd_bit_width, w0_sd_fname, w0_sd_rname)
    rwhw_sd_t = zip(rwhw_sd_bit_width, rwhw_sd_fname, rwhw_sd_rname)
    whw_sd_t = zip(whw_sd_bit_width, whw_sd_fname, whw_sd_rname)
    rw_hw_sd_t = zip(rw_hw_sd_bit_width, rw_hw_sd_fname, rw_hw_sd_rname)

    # WIRE: Register decode and write enable
    for reg in registers:
        wdec_name.append(reg.name.upper())
        wen_name.append(reg.name.upper())

    # WIRE: Register data out
    for reg in registers:
        wdo_name.append(reg.name.upper())

    # assign: address decode
    for reg in registers:
        adec_name.append(reg.name.upper())
        adec_addr_off.append(str(reg.addr_offset).replace("0x", "8'h"))

    adec_t = zip(adec_name, adec_addr_off)

    # assign: write enable
    for reg in registers:
        a_en_rw_name.append(reg.name.upper())

    # Register data out muxing
    for reg in registers:
        dom_name.append(reg.name.upper())
        dom_addr_off.append(str(reg.addr_offset).replace("0x", "32'h"))

    dom_t = zip(dom_name, dom_addr_off)

    for reg in registers:

        # Ungroup data in signals
        reg_in.append(reg_inputs(reg))

        # Instantiation of register cells
        for field in reg.fields:
            if field.access == "rw_hr":
                i_rw_name.append(reg.name.upper())
                i_rw_fbit_w.append(str(field.bit_width))
                i_rw_fname.append(field.name)
                if 0 <= field.reset_val <= 2**field.bit_width-1:
                    i_rw_frstv.append(str(field.reset_val))
                else:
                    print("Reset value does not fit within field ",field.name)
                    exit()
            elif field.access == "rc_w0":
                i_rw0_name.append(reg.name.upper())
                i_rw0_fname.append(field.name)
                i_rw0_fbit_w.append(str(field.bit_width))
                if 0 <= field.reset_val <= 2**field.bit_width-1:
                    i_rw0_frstv.append(str(field.reset_val))
                else:
                    print("Reset value does not fit within field ",field.name)
                    exit()
            elif field.access == "rw_hw":
                i_rw_hw_name.append(reg.name.upper())
                i_rw_hw_fname.append(field.name)
                i_rw_hw_fbit_w.append(str(field.bit_width))
                if 0 <= field.reset_val <= 2**field.bit_width-1:
                    i_rw_hw_frstv.append(str(field.reset_val))
                else:
                    print("Reset value does not fit within field ",field.name)
                    exit()
            elif field.access == "rw_hrw":
                i_rwhw_name.append(reg.name.upper())
                i_rwhw_fname.append(field.name)
                i_rwhw_fbit_w.append(str(field.bit_width))
                if 0 <= field.reset_val <= 2**field.bit_width-1:
                    i_rwhw_frstv.append(str(field.reset_val))
                else:
                    print("Reset value does not fit within field ",field.name)
                    exit()
            elif field.access == "ro":
                i_ro_name.append(reg.name.upper())
                i_ro_fname.append(field.name)
                i_ro_fbit_w.append(str(field.bit_width))
                if 0 <= field.reset_val <= 2**field.bit_width-1:
                    i_ro_frstv.append(str(field.reset_val))
                else:
                    print("Reset value does not fit within field ",field.name)
                    exit()
            elif field.access == "r_hrw":
                i_rhr_name.append(reg.name.upper())
                i_rhr_fname.append(field.name)
                i_rhr_fbit_w.append(str(field.bit_width))
                if 0 <= field.reset_val <= 2**field.bit_width-1:
                    i_rhr_frstv.append(str(field.reset_val))
                else:
                    print("Reset value does not fit within field ",field.name)
                    exit()
            elif field.access == "w_hrw":
                i_whw_name.append(reg.name.upper())
                i_whw_fname.append(field.name)
                i_whw_fbit_w.append(str(field.bit_width))
                if 0 <= field.reset_val <= 2**field.bit_width-1:
                    i_whw_frstv.append(str(field.reset_val))
                else:
                    print("Reset value does not fit within field ",field.name)
                    exit()
            elif field.access == "wo":
                i_wo_name.append(reg.name.upper())
                i_wo_fname.append(field.name)
                i_wo_fbit_w.append(str(field.bit_width))
                if 0 <= field.reset_val <= 2**field.bit_width-1:
                    i_wo_frstv.append(str(field.reset_val))
                else:
                    print("Reset value does not fit within field ",field.name)
                    exit()

        # Group data out signals
        dos_wr_name.append(reg.name.upper())
        dos_wr_out.append(reg_outputs(reg))

        # Connection to the outputs
        for field in reg.fields:
            if field.access == "rw_hr":
                cout_name.append(reg.name.upper())
                cout_fname.append(field.name)
            if field.access == "rw_hrw":
                rwhwcout_name.append(reg.name.upper())
                rwhwcout_fname.append(field.name)
            if field.access == "r_hrw":
                rhrcout_name.append(reg.name.upper())
                rhrcout_fname.append(field.name)
            if field.access == "w_hrw":
                whwcout_name.append(reg.name.upper())
                whwcout_fname.append(field.name)
            if field.access == "wo":
                wocout_name.append(reg.name.upper())
                wocout_fname.append(field.name)

      # # Read error
      # if reg.access != "write-only":
      #     ae_name.append(reg.name.upper())
      #
      # # Write error
      # if reg.access != "read-only":
      #     ae_nro_name.append(reg.name.upper())

    i_rw_t      = zip(i_rw_name, i_rw_fbit_w, i_rw_fname, i_rw_frstv)
    i_rw_t1     = zip(i_rw_name, i_rw_fname)
    i_rw0_t     = zip(i_rw0_name, i_rw0_fbit_w, i_rw0_fname)
    i_rw0_t1    = zip(i_rw0_name, i_rw0_fbit_w, i_rw0_fname, i_rw0_frstv)
    i_rw0_t2    = zip(i_rw0_name, i_rw0_fname)
    i_rw_hw_t   = zip(i_rw_hw_name, i_rw_hw_fbit_w, i_rw_hw_fname)
    i_rw_hw_t1  = zip(i_rw_hw_name, i_rw_hw_fbit_w, i_rw_hw_fname, i_rw_hw_frstv)
    i_rw_hw_t2  = zip(i_rw_hw_name, i_rw_hw_fname)
    i_rwhw_t    = zip(i_rwhw_name, i_rwhw_fbit_w, i_rwhw_fname)
    i_rwhw_t1   = zip(i_rwhw_name, i_rwhw_fbit_w, i_rwhw_fname, i_rwhw_frstv)
    i_rwhw_t2   = zip(i_rwhw_name, i_rwhw_fname)
    i_ro_t      = zip(i_ro_name, i_ro_fname, i_ro_fbit_w, i_ro_frstv)
    i_ro_t1     = zip(i_ro_name, i_ro_fname)
    i_rhr_t     = zip(i_rhr_name, i_rhr_fname, i_rhr_fbit_w, i_rhr_frstv)
    i_whw_t     = zip(i_whw_name, i_whw_fbit_w, i_whw_fname)
    i_whw_t1    = zip(i_whw_name, i_whw_fbit_w, i_whw_fname, i_whw_frstv)
    i_whw_t2    = zip(i_whw_name, i_whw_fname)
    i_wo_t      = zip(i_wo_name, i_wo_fname, i_wo_fbit_w, i_wo_frstv)
    i_wo_t1     = zip(i_wo_name, i_wo_fname)
    dos_wr_t    = zip(dos_wr_name, dos_wr_out)
    cout_t      = zip(cout_name, cout_fname)
    rwhwcout_t  = zip(rwhwcout_name, rwhwcout_fname)
    rhrcout_t   = zip(rhrcout_name, rhrcout_fname)
    whwcout_t   = zip(whwcout_name, whwcout_fname)
    wocout_t    = zip(wocout_name, wocout_fname)

    template = env.get_template("reg_header_template.txt")
    template_r = template.render(module_name=module_name,date=date,out_t=out_t, in_t=in_t, in_w0_t=in_w0_t, in_rwhw=in_rwhw, in_rhr=in_rhr, in_hwh=in_hwh, in_wo=in_wo,
                                 res_t=res_t, inf_t=inf_t, ino_t=ino_t, en_t=en_t, w0_sd_t=w0_sd_t, rwhw_sd_t=rwhw_sd_t, whw_sd_t=whw_sd_t,
                                 wdec_name=wdec_name, wen_name=wen_name, wdo_name=wdo_name, irw_hw_t=irw_hw_t,
                                 adec_t=adec_t, a_en_rw_name=a_en_rw_name, i_wo_t=i_wo_t, i_wo_t1=i_wo_t1, rw_hw_sd_t=rw_hw_sd_t,
                                 dom_t=dom_t, reg_in=reg_in, i_rw_t=i_rw_t, i_rw_t1=i_rw_t1, i_rw0_t=i_rw0_t, i_rw0_t1=i_rw0_t1, i_rw_hw_t=i_rw_hw_t, i_rw_hw_t1=i_rw_hw_t1, i_rw_hw_t2=i_rw_hw_t2,
                                 i_ro_t=i_ro_t, i_ro_t1=i_ro_t1, i_rhr_t=i_rhr_t, i_whw_t=i_whw_t, i_whw_t1=i_whw_t1, i_whw_t2=i_whw_t2,  i_rw0_t2=i_rw0_t2, i_rwhw_t=i_rwhw_t, i_rwhw_t1=i_rwhw_t1, i_rwhw_t2=i_rwhw_t2, dos_wr_t=dos_wr_t,
                                 cout_t=cout_t, rwhwcout_t=rwhwcout_t, rhrcout_t=rhrcout_t, whwcout_t=whwcout_t, wocout_t=wocout_t, ae_name=ae_name, ae_nro_name=ae_nro_name)

    if os.path.exists(module_file):
        os.remove(module_file)

    # Copy the content of the template into the .sv file.
    with open(module_file, 'a') as file_handler:
        file_handler.write(template_r)

    # Replacing the last comma with ");" at the end of the interface.
    with open(module_file, 'r') as file:
        content = file.read()
    index_of_comma = content.rfind(',', 0, content.find('WIRE_definition'))
    if index_of_comma != -1:
        updated_content = content[:index_of_comma] + ');' + content[index_of_comma + 1:]
        with open(module_file, 'w') as file:
            file.write(updated_content)
    else:
        print("The comma before 'WIRE_definition' was not found.")
    with open(module_file, 'r') as file:
        lines = file.readlines()
    wire_definition_index = None
    for i, line in enumerate(lines):
        if '/*********WIRE_definition*********/' in line:
            wire_definition_index = i
            break
    if wire_definition_index is not None:
        while wire_definition_index > 0 and lines[wire_definition_index - 1].strip() == '':
            del lines[wire_definition_index - 1]
            wire_definition_index -= 1
    with open(module_file, 'w') as file:
        file.writelines(lines)

    # Successfully generated REGFILE.
    print("Registers successfully generated.")
    script_path = os.path.dirname(os.path.abspath(__file__))
    print("The register file is located at the following path:", script_path)
    print("Module name:", module_name+".sv")
    print("Register names:")
    for x in registers:
        print(x.name)

